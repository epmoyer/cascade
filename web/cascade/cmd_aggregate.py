"""Handler for the 'aggregate' command
(Commands are issued on the command line, per the docopt syntax in __main__.py)
"""

# Standard library
import os
import zipfile
import tempfile
import shutil
import json

# Libraries
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter

# Local
from cascade import cmd_check
from cascade.word_docx import WordDocx
from cascade import quicklog
from cascade.util_eliot import log_function

qlog = quicklog.get_logger()
lprint = qlog.lprint

MAX_XLSX_COL_WIDTH = 50

@log_function
def aggregate(arguments):
    """Create aggregation summary from multiple requirements documents

    * Receive a zip containing multiple requirements docs
    * Parse all to JSON
    * Create an output summary spreadsheet (.xlsx)
        * Aggregate each requirement into a row
        * Columnize the enumerated type (“allocatedTo”) field into multiple columns
            * Make each a filterable checkbox (“X”) list
        * Include src doc as a column field.

    Returns:
        Tuple of output filenames if successful
        None otherwise
    """
    zip_directory_in = None
    staging_directory_out = None
    output_xlsx_filename = 'aggregation.xlsx'

    #-----------------------------
    # Extract submitted files from .zip
    #-----------------------------
    # Note: For this operation, the submitted file will be a .zip file (not a
    #       .docx file as the argument name implies)
    zip_filename_in = arguments['<requirements.docx>']
    if not os.path.isfile(zip_filename_in):
        qlog.error(f'The file "{zip_filename_in}" does not exist')
        return False
    if not zip_filename_in.endswith('.zip'):
        qlog.error(f'Expected filename ("{zip_filename_in}") to end in ".zip".')
        return False

    lprint(f'Extracting files from "{zip_filename_in}"...')
    if not os.path.exists('temp'):
        os.makedirs('temp')
    zip_directory_in = tempfile.mkdtemp(prefix='zip_extract_', dir='temp')
    zip_ref = zipfile.ZipFile(zip_filename_in, 'r')
    zip_ref.extractall(zip_directory_in)
    zip_ref.close()

    #-----------------------------
    # Find .docx files in .zip
    #-----------------------------
    files = [
        f for f in os.listdir(zip_directory_in)
        if os.path.isfile(os.path.join(zip_directory_in, f))
        ]
    lprint('Found files in .zip:\n' + '\n'.join([f'    "{f}"' for f in files]))
    docx_files = [f for f in files if f.endswith(".docx")]
    if not docx_files:
        qlog.error(f'Expected zip file to contain at least one file ending in ".docx"')
        return False

    #-----------------------------
    # Extract requirements from each requirements .docx file
    #-----------------------------
    staging_directory_out = tempfile.mkdtemp(prefix='staging_', dir='temp')
    document_dicts = []
    for filename in docx_files:
        docx_filename = os.path.join(zip_directory_in, filename)

        #-----------------------------
        # Check integrity
        #-----------------------------
        check_arguments = {'<requirements.docx>':docx_filename}
        if not cmd_check.check(check_arguments):
            qlog.error('Operation aborted due to document check failures.')
            return False

        #-----------------------------
        # Extract requirement data
        #-----------------------------
        lprint('Extracting requirements from "{}"...'.format(docx_filename))
        doc = WordDocx(qlog, os.path.abspath(docx_filename))
        document_dict = {
            'sourceDocument': filename,
            'directives': [
                directive.as_dict
                for directive in doc._directives
                if '#document_info' not in directive.as_dict
                ]
        }
        qlog.debug(f'document_dict: {json.dumps(document_dict, indent=4)}')
        document_dicts.append(document_dict)

        out_filename = os.path.join(staging_directory_out, filename.replace('.docx', '.json'))
        with open(out_filename, "w") as text_file:
            text_file.write(json.dumps(document_dict, indent=4))

    #-----------------------------
    # Aggregate directives into single .xlsx
    #-----------------------------
    exporter = ExcelExporter(os.path.join(staging_directory_out, output_xlsx_filename))
    exporter.add_column_header('id')
    # Note: For now the list of directive properties which are enumerations is
        #       hard coded. It may be possible to be more general in the
        #       future, and discover which directive properties are enumerations by
        #       parsing the "#document_info" directive in each document (ant presumably
        #       identifying which ones have a type of "array" in their schema)
    exporter.add_ennumeration(
        'allocatedTo',
        ['System', 'Application', 'Function', 'Network', 'Equipment', 'Radio']
        )
    for document_dict in document_dicts:
        for directive in document_dict['directives']:
            directive['sourceDocument'] = document_dict['sourceDocument']
            exporter.add_row(directive)
    exporter.prettify()
    exporter.save()

    #-----------------------------
    # Return the output files
    #-----------------------------
    return_directory = os.path.join('cascade', 'static', 'results')
    return_filename_list = []
    filenames = [
        f for f in os.listdir(staging_directory_out)
        if os.path.isfile(os.path.join(staging_directory_out, f))
        ]
    for filename in filenames:
        source = os.path.join(staging_directory_out, filename)
        destination = os.path.join(return_directory, filename)
        shutil.move(source, destination)
        return_filename_list.append(filename)
    # Move primary output file to end of list
    return_filename_list.append(
        return_filename_list.pop(
            return_filename_list.index(
                output_xlsx_filename)))

    #-----------------------------
    # Delete temp directories
    #-----------------------------
    for directory in (zip_directory_in, staging_directory_out):
        if directory is not None:
            lprint('Deleting temp directory "{}"'.format(directory))
            try:
                # qlog.critical('Temp dir deletion stubbed out.')
                # pass
                shutil.rmtree(directory)
            except Exception as e:
                qlog.error('Directory deletion failed.\n   {}'.format(str(e)))

    lprint("Done.")
    return return_filename_list

class ExcelExporter():
    """ Export requirement directives to an Excel (.xlsx) file """

    def __init__(self, filename):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.filename = filename

        self.column_map = {}
        self.tier_1_header_row = 1
        self.tier_2_header_row = self.tier_1_header_row + 1
        self.current_row = self.tier_2_header_row + 1
        self.next_column = 1

        self.enumeration_properties = []

        self.fill = PatternFill(
            patternType='solid',
            fill_type='solid',
            fgColor=Color('EBF1DE') # RGB: 235, 241, 222
            )
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
            )
        self.thin_border_left = Border(
            left=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
            )
        self.thin_border_right = Border(
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
            )
        self.thin_border_center = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin')
            )

    def add_row(self, data_dict):
        """Add a row containing all properties from a dict of the requirements directive"""
        for key, value in data_dict.items():
            if key in self.enumeration_properties:
                for value2 in value:
                    self.add_property(value2, 'X')      
            else:
                if isinstance(value, (list,)):
                    self.add_property(key, ', '.join(value))
                else:
                    self.add_property(key, value)
        self.current_row += 1

    def add_property(self, key, value):
        """add a new property to the current row"""
        self.add_column_header(key)
        self.ws.cell(row=self.current_row, column=self.column_map[key], value=value)

    def add_column_header(self, key):
        """Add a column header if it doesn't already exist"""
        if key not in self.column_map:
            self.column_map[key] = self.next_column
            self.ws.cell(row=self.tier_2_header_row, column=self.next_column, value=key)
            self.next_column += 1

    def add_ennumeration(self, name, values):
        """Add columns for an enumeration

        The enumeration name shows up on the top row.
        The enumeration values show up on the primary heading row.
        """
        start_column = self.next_column
        self.enumeration_properties.append(name)
        cell = self.ws.cell(row=self.tier_1_header_row, column=self.next_column, value=name)
        for value in values:
            self.add_column_header(value)
        for column in range(start_column, self.next_column):
            cell = self.ws.cell(row=self.tier_1_header_row, column=column)
            if column == start_column:
                cell.border = self.thin_border_left
            elif column == self.next_column-1:
                cell.border = self.thin_border_right
            else:
                cell.border = self.thin_border_center
            cell.fill = self.fill

    def save(self):
        """Save the current workbook"""
        self.wb.save(self.filename)

    def prettify(self):
        """Format all worksheets

        * Add fill color to heading columns.
        * Set column widths based on content size
        * Set border lines
        """
        worksheet_names = self.wb.get_sheet_names()
        alignment_default = Alignment(horizontal='left', wrapText=True, vertical='top')
        for table_name in worksheet_names:
            sheet = self.wb.get_sheet_by_name(table_name)
            sheet.column_dimensions['A'].width = 41
            for cell in sheet['A1':'A200']:
                cell[0].alignment = alignment_default

        for worksheet_name in worksheet_names:
            worksheet = self.wb.get_sheet_by_name(name=worksheet_name)
            dims = {}

            end_reached = False
            max_row = 0
            max_col = 0
            for row_index, row in enumerate(worksheet.rows):
                if row_index == self.tier_1_header_row-1:
                    continue
                for col_index, cell in enumerate(row):
                    max_col = max(max_col, col_index)
                    if row_index > 0 and col_index == 0 and not cell.value:
                        # Stop on fist blank row heading
                        end_reached = True
                        max_row = row_index - 1
                    if not end_reached:
                        if row_index == self.tier_2_header_row - 1:
                            # Column headings
                            width_scale = 1.4
                        else:
                            # Content cells
                            width_scale = 1.2
                        if cell.value:
                            column_width = len(str(cell.value)) * width_scale
                            if col_index > 0:
                                column_width = min(column_width, MAX_XLSX_COL_WIDTH)
                            dims[cell.column] = max((dims.get(cell.column, 0), column_width))
                        if row_index == self.tier_2_header_row - 1:
                            cell.fill = self.fill
                        cell.border = self.thin_border
                        if col_index > 0:
                            cell.alignment = alignment_default

            worksheet.auto_filter.ref = "B2:{}{}".format(get_column_letter(max_col + 1), max_row)

            for col, value in dims.items():
                worksheet.column_dimensions[col].width = value
