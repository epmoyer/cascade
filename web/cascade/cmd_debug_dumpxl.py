"""Handler for the 'debug-dumpxl' command
(Commands are issued on the command line, per the docopt syntax in __main__.py)
"""

# Standard library
import os

# Libraries
from openpyxl import load_workbook

# Local
from cascade import quicklog
from cascade.util_eliot import log_function

qlog = quicklog.get_logger()
lprint = qlog.lprint

@log_function
def dump(arguments):
    '''Dump the details of an Excel .xlsx (or .xlsm) file for development debugging'''

    filename = arguments['<CrdFeatureInfo.xlsm>']

    if not os.path.isfile(filename):
        qlog.error('The file "{}" does not exist'.format(filename))
        return

    import warnings
    warnings.simplefilter("ignore")
    workbook = load_workbook(
        filename=filename,
        data_only=True  # Reads will return cell values, NOT formulas
        )
    warnings.simplefilter("default")

    sheet_names = workbook.sheetnames
    print('Sheet names: {}'.format(sheet_names))

    worksheet = workbook['Features']

    col = 'N'
    for row in (1, 2, 3, 4, 5, 6, 7, 37, 38, 98, 40, 41):
        cell_id = col + str(row)
        print('   {}: {}: {}'.format(cell_id, type(worksheet[cell_id].value), worksheet[cell_id].value))
    #worksheet[cell_loc(row_number, col_number)].value
